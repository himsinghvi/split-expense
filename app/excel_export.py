"""Build multi-sheet Excel expense reports for an event."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import services
from app.models import OrganizationMember

HEADER_FILL = PatternFill("solid", fgColor="E7E5E4")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
META_FONT = Font(size=11)
THIN = Side(style="thin", color="CCC8C5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _safe_filename(name: str, event_id: int) -> str:
    base = re.sub(r"[^\w\s-]", "", name, flags=re.UNICODE) or "event"
    base = re.sub(r"[-\s]+", "_", base).strip("_")[:50]
    d = datetime.now(timezone.utc).strftime("%Y%m%d")
    return f"expense_report_{base}_{event_id}_{d}.xlsx"


def _money_num(v) -> float:
    return float(Decimal(str(v)).quantize(Decimal("0.01")))


def _excel_datetime(dt: datetime) -> datetime:
    """openpyxl rejects timezone-aware datetimes; store as naive UTC wall time."""
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def _style_header_row(ws, row: int, col_start: int, col_end: int) -> None:
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, col_start: int, col_end: int, min_w: float = 10, max_w: float = 48) -> None:
    for col in range(col_start, col_end + 1):
        letter = get_column_letter(col)
        maxlen = min_w
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            maxlen = max(maxlen, min(max_w, len(str(v)) + 2))
        ws.column_dimensions[letter].width = maxlen


def build_event_report_xlsx(db: Session, event_id: int) -> tuple[bytes, str] | None:
    """Return (xlsx bytes, filename) or None if event missing."""
    ev = services.get_event_for_report(db, event_id)
    if not ev:
        return None

    org_name = ev.organization.name if ev.organization else ""
    org_id = ev.organization_id
    event_balances = services.member_balances(db, event_id)
    org_balances = services.org_member_balances(db, org_id)
    oms = list(
        db.scalars(
            select(OrganizationMember).where(OrganizationMember.organization_id == org_id)
        ).all()
    )
    om_id_to_uid = {om.id: om.user_id for om in oms}
    uid_to_org: dict[int, object] = {}
    for b in org_balances:
        uid = om_id_to_uid.get(b.member_id)
        if uid is not None:
            uid_to_org[int(uid)] = b

    expenses = sorted(
        ev.expenses, key=lambda e: (e.expense_date, e.id), reverse=True
    )

    org_full = services.get_organization(db, org_id)
    contrib_rows: list[tuple[datetime, str, float, str]] = []
    if org_full:
        for c in sorted(
            org_full.org_contributions or [],
            key=lambda x: x.created_at,
            reverse=True,
        ):
            u = c.user
            label = ((u.full_name or "").strip() or u.mobile) if u else str(c.user_id)
            contrib_rows.append(
                (
                    c.created_at,
                    label,
                    _money_num(c.amount),
                    (c.note or "").strip(),
                )
            )

    wb = Workbook()
    ws_sum = wb.active
    assert ws_sum is not None
    ws_sum.title = "Summary"

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws_sum["A1"] = "Expense report"
    ws_sum["A1"].font = TITLE_FONT
    ws_sum["A2"] = f"Organization: {org_name}"
    ws_sum["A2"].font = META_FONT
    ws_sum["A3"] = f"Event: {ev.name}"
    ws_sum["A3"].font = META_FONT
    ws_sum["A4"] = f"Generated: {now}"
    ws_sum["A4"].font = META_FONT
    ws_sum["A5"] = f"Event ID: {ev.id}"
    ws_sum["A5"].font = META_FONT

    hdr_row = 7
    pool_avail = services.org_pool_available(db, org_id)
    ws_sum.cell(row=hdr_row, column=1, value="Organization pool (unspent)")
    ws_sum.cell(row=hdr_row, column=2, value=float(pool_avail))
    ws_sum.cell(row=hdr_row, column=1).font = META_FONT
    ws_sum.cell(row=hdr_row, column=2).font = META_FONT
    hdr_row += 2

    headers_bal = [
        "Member",
        "This event spent",
        "Org pooled",
        "Org expended",
        "Org remaining",
    ]
    for i, h in enumerate(headers_bal, start=1):
        ws_sum.cell(row=hdr_row, column=i, value=h)
    _style_header_row(ws_sum, hdr_row, 1, len(headers_bal))

    r = hdr_row + 1
    for b in event_balances:
        m = next((mm for mm in ev.members if mm.id == b.member_id), None)
        ob = uid_to_org.get(m.user_id) if m and m.user_id else None
        ws_sum.cell(row=r, column=1, value=b.name)
        ws_sum.cell(row=r, column=2, value=_money_num(b.expended))
        ws_sum.cell(row=r, column=3, value=_money_num(ob.contributed) if ob else 0.0)
        ws_sum.cell(row=r, column=4, value=_money_num(ob.expended) if ob else 0.0)
        ws_sum.cell(row=r, column=5, value=_money_num(ob.remaining) if ob else 0.0)
        for c in range(1, 6):
            ws_sum.cell(row=r, column=c).border = BORDER
        r += 1

    if event_balances:
        ws_sum.auto_filter.ref = (
            f"A{hdr_row}:{get_column_letter(len(headers_bal))}{hdr_row + len(event_balances)}"
        )
    _auto_width(ws_sum, 1, 5)

    ws_exp = wb.create_sheet("Expenses")
    exp_headers = ["Date", "Item", "Category", "Total", "Split summary"]
    for i, h in enumerate(exp_headers, start=1):
        ws_exp.cell(row=1, column=i, value=h)
    _style_header_row(ws_exp, 1, 1, len(exp_headers))

    for ri, e in enumerate(expenses, start=2):
        parts = []
        for s in sorted(e.splits, key=lambda x: x.member.name):
            parts.append(f"{s.member.name}: {_money_num(s.amount):.2f}")
        split_txt = "; ".join(parts)
        ws_exp.cell(row=ri, column=1, value=e.expense_date)
        ws_exp.cell(row=ri, column=2, value=e.title)
        ws_exp.cell(row=ri, column=3, value=e.category)
        ws_exp.cell(row=ri, column=4, value=_money_num(e.amount_total))
        ws_exp.cell(row=ri, column=5, value=split_txt)
        for c in range(1, 6):
            ws_exp.cell(row=ri, column=c).border = BORDER

    if expenses:
        ws_exp.auto_filter.ref = f"A1:{get_column_letter(len(exp_headers))}{1 + len(expenses)}"
    _auto_width(ws_exp, 1, 5)

    ws_c = wb.create_sheet("Org pool")
    ch = ["When (UTC)", "Person", "Amount", "Note"]
    for i, h in enumerate(ch, start=1):
        ws_c.cell(row=1, column=i, value=h)
    _style_header_row(ws_c, 1, 1, len(ch))

    for ri, (ts, mname, amt, note) in enumerate(contrib_rows, start=2):
        ws_c.cell(row=ri, column=1, value=_excel_datetime(ts))
        ws_c.cell(row=ri, column=2, value=mname)
        ws_c.cell(row=ri, column=3, value=amt)
        ws_c.cell(row=ri, column=4, value=note or "")
        for c in range(1, 5):
            ws_c.cell(row=ri, column=c).border = BORDER

    if contrib_rows:
        ws_c.auto_filter.ref = f"A1:{get_column_letter(len(ch))}{1 + len(contrib_rows)}"
    _auto_width(ws_c, 1, 4)

    ws_s = wb.create_sheet("Split lines")
    sh = [
        "Expense date",
        "Item",
        "Category",
        "Expense total",
        "Member",
        "Split amount",
    ]
    for i, h in enumerate(sh, start=1):
        ws_s.cell(row=1, column=i, value=h)
    _style_header_row(ws_s, 1, 1, len(sh))

    ri = 2
    for e in sorted(ev.expenses, key=lambda x: (x.expense_date, x.id)):
        for s in sorted(e.splits, key=lambda x: x.member.name):
            ws_s.cell(row=ri, column=1, value=e.expense_date)
            ws_s.cell(row=ri, column=2, value=e.title)
            ws_s.cell(row=ri, column=3, value=e.category)
            ws_s.cell(row=ri, column=4, value=_money_num(e.amount_total))
            ws_s.cell(row=ri, column=5, value=s.member.name)
            ws_s.cell(row=ri, column=6, value=_money_num(s.amount))
            for c in range(1, 7):
                ws_s.cell(row=ri, column=c).border = BORDER
            ri += 1

    if ri > 2:
        ws_s.auto_filter.ref = f"A1:{get_column_letter(len(sh))}{ri - 1}"
    _auto_width(ws_s, 1, 6)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = _safe_filename(ev.name, ev.id)
    return bio.getvalue(), filename
